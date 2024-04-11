import os, io
from io import BytesIO
import sys
import numpy as np
import pandas as pd
from flask import request, render_template, send_file, send_from_directory
import plotly.graph_objs as go
from plotly.offline import plot
from plotly.subplots import make_subplots
import path 
import yfinance as yf
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import tempfile
import shutil

# directory reach
directory = path.Path(__file__).abspath()
 
# setting path
sys.path.append(directory.parent.parent)

import models.lending_value as lev

def clear_temp_directory():
    temp_dir = os.path.join('src', 'app', 'temp_files')
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
        os.makedirs(temp_dir)  # Recreate the directory after clearing

def calculate_avg_close_price(ticker):
    ticker = ticker.strip().upper() + '.SW'
    try:
        stock_data = yf.download(ticker, period="1mo", interval="1d")
        if len(stock_data) < 2:
            return None
        avg_close_price = stock_data['Close'].mean()
        return round(avg_close_price, 2)
    except Exception as e:
        print(f"Error processing {ticker}: {e}")
        return None
    
def calculate_volatility(ticker):
    ticker = ticker.strip().upper() + '.SW'
    try:
        stock_data = yf.download(ticker, period="1mo", interval="1d")
        if len(stock_data) < 2:
            return None
        daily_returns = stock_data['Close'].pct_change()
        volatility = daily_returns.std() * np.sqrt(252)
        return volatility
    except Exception as e:
        print(f"Error processing {ticker}: {e}")
        return None

def calculate_adtv(ticker):
    ticker = ticker.strip().upper() + '.SW'
    try:
        stock_data = yf.download(ticker, period="1mo", interval="1d")
        if len(stock_data) < 2:
            return None
        adtv = stock_data['Volume'].mean()
        return round(adtv)
    except Exception as e:
        print(f"Error processing {ticker}: {e}")
        return None

def update_excel(workbook):
    start_row=2
    ticker_column_index = 1
    vola_column_index = 5
    adtv_column_index = 4
    avg_close_column_index = 3

    sheet = workbook.active

    for row in range(start_row, sheet.max_row + 1):
        ticker = sheet.cell(row=row, column=ticker_column_index).value
        if ticker:
            volatility = calculate_volatility(ticker)
            adtv = calculate_adtv(ticker)
            avg_close_price = calculate_avg_close_price(ticker)

            if volatility is not None:
                sheet.cell(row=row, column=vola_column_index, value=volatility)
            if adtv is not None:
                sheet.cell(row=row, column=adtv_column_index, value=adtv)
            if avg_close_price is not None:
                sheet.cell(row=row, column=avg_close_column_index, value=avg_close_price)

def calc_ptf():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'

        file = request.files['file']

        if file.filename == '':
            return 'No selected file'

        if file:
            # read the file into a BytesIO buffer
            input_buffer = BytesIO(file.read())
            
            # load the workbook from the buffer
            workbook = openpyxl.load_workbook(filename=BytesIO(input_buffer.getvalue()))
            update_excel(workbook)

            # save the updated workbook to a new BytesIO buffer for download
            output_buffer = BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)

            download_name = "processed_" + file.filename

            return send_file(output_buffer, as_attachment=True, download_name=download_name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('portfolio_builder.html')

def single_plot(sigma, delta, epsilon, alpha, size, adtv, price):
    x_values = np.logspace(0, np.maximum(round(len(str(adtv))*1.5, None), round(len(str(size))*1.5, None)), 100, base=10)
    x_values = np.append(x_values, size)
    x_values = np.unique(x_values)
    lv_values = []
    loan_values = []
    mu = (sigma ** 2) / 2
    gamma = 10 ** -1.87096 * adtv ** -0.794554

    for x in x_values:
        if x == size:
            size_lv = lev.calc_lv(mu, sigma, delta, epsilon, alpha, size, gamma)
            lv_values.append(size_lv)
            size_loan = size_lv * size * price
            loan_values.append(size_loan)
        else:
            lv = lev.calc_lv(mu, sigma, delta, epsilon, alpha, x, gamma)
            lv_values.append(lv)
            loan = lv * x * price
            loan_values.append(loan)

    tickvals = [10**i for i in range(0, np.maximum(round(len(str(adtv))*1.5, None), round(len(str(size))*1.5, None)))]

    # create subplots
    fig = make_subplots(rows=2, cols=1)

    # add trace
    fig.add_trace(
        go.Scatter(x=x_values, y=lv_values, mode='lines', name='LV', marker=dict(color='#015687'),
                   hovertemplate='# of stocks: %{x:,.0f}<br>LV: %{y:.2%}<extra></extra>'),
        row=1, col=1
    )
    fig.add_trace(
        go.Scatter(x=[size], y=[size_lv], mode='markers', name='Input', marker=dict(color='red'),
                   hovertemplate='# of stocks: %{x:,.0f}<br>LV: %{y:.2%}<extra></extra>'),
        row=1, col=1
    )
    fig.add_trace(
        go.Scatter(x=x_values, y=loan_values, mode='lines', name='Loan Amount', marker=dict(color='#00C1FF'),
                   hovertemplate='# of stocks: %{x:,.0f}<br>Loan Amount: %{y:,.2f} CHF<extra></extra>'),
        row=2, col=1
    )
    fig.add_trace(
        go.Scatter(x=[size], y=[size_loan], mode='markers', name='Input', marker=dict(color='red'),
                    hovertemplate='# of stocks: %{x:,.0f}<br>Loan Amount: %{y:,.2f} CHF<extra></extra>'),
        row=2, col=1
    )

    # layout
    fig.update_layout(
        xaxis=dict(
            type='log',
            tickvals=tickvals
        ),
        xaxis2=dict(
            type='log',
            tickvals=tickvals,
            title='No. of Stocks'
        ),
        yaxis=dict(
            title='Liquidity-adjusted LV',
            tickvals=np.arange(0, 1.2, 0.2)
        ),
        yaxis2=dict(
            title='Loan Amount (CHF)'
        ),
        showlegend=False,
        height = 500,
        margin=dict(l=80, r=80, t=20, b=60)
    )

    plot_div = plot(fig, output_type='div', include_plotlyjs=False)

    return plot_div

def multi_plot(df, delta, epsilon, alpha):
    fig = make_subplots(rows=2, cols=1)

    colors = ['#015687', '#00C1FF', '#FF5733', '#33FF57', '#5733FF', '#FFC300', '#DAF7A6', '#581845', '#C70039', '#900C3F']
    line_styles = ['solid', 'dot', 'dash', 'longdash', 'dashdot', 'longdashdot']

    style_combinations = [(color, 'solid') for color in colors]
    style_combinations += [(color, style) for color in colors for style in line_styles if style != 'solid']
    x_scale = np.maximum(round(len(str(df['ADTV'].max()))*1.5, None), round(len(str(df['Shares'].max()))*1.5, None))
    tickvals = [10**i for i in range(0, x_scale)]

    for index, (i, row) in enumerate(df.iterrows()):
        color, style = style_combinations[index % len(style_combinations)]

        ticker = row['Symbol']
        sigma = row['Vola']
        x = row['Shares']
        price = row['Price (CHF)']
        mu = (row['Vola'] ** 2) / 2
        gamma = 10 ** -1.87096 * row['ADTV'] ** -0.794554

        x_values = np.logspace(0, x_scale, 100, base=10)
        lv_values = [lev.calc_lv(mu, sigma, delta, epsilon, alpha, val, gamma) for val in x_values]
        loan_values = [lv_val * val * price for lv_val, val in zip(lv_values, x_values)]
        size_lv = lev.calc_lv(mu, sigma, delta, epsilon, alpha, x, gamma)
        size_loan = size_lv * x * price

        # add trace
        fig.add_trace(
            go.Scatter(x=x_values, y=lv_values, mode='lines', legendgroup=ticker, name=ticker, line=dict(color=color, dash=style),
                    hovertemplate='# of stocks: %{x:,.0f}<br>LV: %{y:.2%}<extra></extra>', showlegend=True),
                    row=1, col=1
                )
        fig.add_trace(
            go.Scatter(x=[x], y=[size_lv], mode='markers', legendgroup=ticker, marker=dict(color=color),
                    hovertemplate='# of stocks: %{x:,.0f}<br>LV: %{y:.2%}<extra></extra>', showlegend=False),
            row=1, col=1
        )
        fig.add_trace(
            go.Scatter(x=x_values, y=loan_values, mode='lines', legendgroup=ticker, line=dict(color=color, dash=style),
                    hovertemplate='# of stocks: %{x:,.0f}<br>Loan Amount: %{y:,.2f} CHF<extra></extra>', showlegend=False),
            row=2, col=1
        )
        fig.add_trace(
            go.Scatter(x=[x], y=[size_loan], mode='markers', legendgroup=ticker, marker=dict(color=color),
                        hovertemplate='# of stocks: %{x:,.0f}<br>Loan Amount: %{y:,.2f} CHF<extra></extra>', showlegend=False),
            row=2, col=1
        )
    
    fig.update_layout(
        xaxis=dict(
            type='log',
            tickvals=tickvals
        ),
        yaxis=dict(
            title='Liquidity-adjusted LV',
            tickvals=np.arange(0, 1.2, 0.2)
        ),
        xaxis2=dict(
            type='log',
            tickvals=tickvals,
            title='No. of Stocks',
            anchor='y2'
        ),
        yaxis2=dict(
            title='Loan Amount (CHF)',
            anchor='x2'
        ),
        showlegend=True,
        height = 500,
        margin=dict(l=80, r=80, t=20, b=60)
    )

    plot_div = plot(fig, output_type='div', include_plotlyjs=False)
    return plot_div

def table_out(df, delta, epsilon, alpha):
    swx_list = pd.read_csv(os.path.join('data', 'external', 'swx_symbol_list.csv'))
    tick_list = swx_list['Symbol']
    expected_columns = {'Symbol', 'Shares', 'Price (CHF)', 'ADTV', 'Vola'}
    df['Symbol'] = df['Symbol'].astype('str').str.upper()

    if not expected_columns.issubset(df.columns):
        raise ValueError('Missing one or more required columns.')
    
    # filter out rows with empty values
    invalid_df_empty = df[df.isnull().any(axis=1)].copy()
    df = df.dropna().copy()

    invalid_df_list = df[~df['Symbol'].isin(tick_list)].copy()  # rows not in the list
    df = df[df['Symbol'].isin(tick_list)].copy()  # rows in the list

    # combine invalid dataframes
    invalid_df = pd.concat([invalid_df_empty, invalid_df_list]).drop_duplicates(subset=df.index.name).sort_index()
    invalid_df.index += 2

    df['Mu'] = (df['Vola'] ** 2) / 2
    df['Gamma'] = 10 ** -1.87096 * df['ADTV'] ** -0.794554
    df['LV'] = df.apply(lambda row: lev.calc_lv(row['Mu'], row['Vola'], delta, epsilon, alpha, 0, row['Gamma']), axis=1)
    df['Liq LV'] = df.apply(lambda row: lev.calc_lv(row['Mu'], row['Vola'], delta, epsilon, alpha, row['Shares'], row['Gamma']), axis=1)
    df['Loan Amount (CHF)'] = df['Shares'] * df['Liq LV'] * df['Price (CHF)']
    total_loan = df['Loan Amount (CHF)'].sum()

    out_df = df.copy()
    out_df['LV'] = out_df['LV'].map('{:.2%}'.format)
    out_df['Liq LV'] = out_df['Liq LV'].map('{:.2%}'.format)
    out_df['Loan Amount (CHF)'] = out_df['Loan Amount (CHF)'].map('{:,.2f}'.format)

    out_table_html = out_df.filter(['Symbol', 'Shares', 'Price (CHF)', 'ADTV', 'Vola', 'LV', 'Liq LV', 'Loan Amount (CHF)'], axis=1).to_html(classes='data-table', header="true", index=False)
    invalid_table_html = invalid_df.copy().to_html(classes='data-table', header="true", index=True)

    if invalid_df.empty:
        invalid_message = ''
    else:
        invalid_message = 'Invalid entries:'

    return df, total_loan, out_table_html, invalid_table_html, out_df, invalid_df, invalid_message

def multi_asset():
    # default values and init
    plot_div = ''
    default_delta = 10 / 250
    default_epsilon = 0.01
    default_alpha = 0.25
    delta, epsilon, alpha, total_loan = default_delta, default_epsilon, default_alpha, 0
    error_message, invalid_message, table_html, invalid_table_html = '', '', '', ''
    out_df, invalid_df, file_processed = pd.DataFrame(), pd.DataFrame(), False

    if request.method == 'POST':
        if 'download' in request.form:
            filename = request.form.get('filename')
            return send_from_directory('temp_files', filename, as_attachment=True)
        clear_temp_directory()
        try:
            # inputs
            delta = float(request.form.get('delta', default_delta))
            epsilon = float(request.form.get('epsilon', default_epsilon))
            alpha = float(request.form.get('alpha', default_alpha))

            # input errors
            if not 0 <= delta <= 1 or not 0 <= epsilon <= 1 or not 0 <= alpha <= 1:
                raise ValueError('Epsilon, Alpha, and Delta must be between 0 and 1.')

            file = request.files.get('file')
            if file:
                df = pd.read_excel(file, usecols=['Symbol', 'Shares', 'Price (CHF)', 'ADTV', 'Vola'])
                df, total_loan, table_html, invalid_table_html, out_df, invalid_df, invalid_message = table_out(df, delta, epsilon, alpha)
                plot_div = multi_plot(df, delta, epsilon, alpha)
                total_loan = format(total_loan, ',.2f')
                file_processed = True
                output = io.BytesIO()
                workbook = openpyxl.Workbook()
                out_sheet = workbook.active
                out_sheet.title = 'Output'
                
                # write DataFrame to Excel sheet
                for row in dataframe_to_rows(out_df, index=False, header=True):
                    out_sheet.append(row)
                
                invalid_sheet = workbook.create_sheet(title='Invalid Entries')
                for row in dataframe_to_rows(invalid_df, index=True, header=True):
                    invalid_sheet.append(row)
                
                workbook.save(output)
                output.seek(0)
                temp_dir = os.path.join('src', 'app', 'temp_files')
                os.makedirs(temp_dir, exist_ok=True)
                temp_filename = tempfile.NamedTemporaryFile(delete=True, dir=temp_dir, suffix='.xlsx').name
                with open(temp_filename, 'wb') as f:
                    f.write(output.getvalue())

                generated_filename = os.path.basename(temp_filename)

                return render_template('multi_asset.html', 
                            file_processed=file_processed,
                            plot_div=plot_div, 
                            error_message=error_message, 
                            invalid_message=invalid_message, 
                            table_html=table_html, 
                            invalid_table_html=invalid_table_html, 
                            delta=delta, epsilon=epsilon, alpha=alpha,
                            default_delta=default_delta, 
                            default_epsilon=default_epsilon, 
                            default_alpha=default_alpha, 
                            total_loan=total_loan,
                            generated_filename=generated_filename)
            
        except ValueError as e:     
            error_message = f'<b style="color: red;">WARNING:</b> {e}'

        return render_template('multi_asset.html', 
                        file_processed=file_processed,
                        plot_div=plot_div, 
                        error_message=error_message, 
                        invalid_message=invalid_message, 
                        table_html=table_html, 
                        invalid_table_html=invalid_table_html, 
                        delta=delta, epsilon=epsilon, alpha=alpha,
                        default_delta=default_delta, 
                        default_epsilon=default_epsilon, 
                        default_alpha=default_alpha, 
                        total_loan=total_loan)
    else:
        return render_template('multi_asset.html', 
                        file_processed=file_processed,
                        plot_div=plot_div, 
                        error_message=error_message, 
                        invalid_message=invalid_message, 
                        table_html=table_html, 
                        invalid_table_html=invalid_table_html, 
                        delta=delta, epsilon=epsilon, alpha=alpha,
                        default_delta=default_delta, 
                        default_epsilon=default_epsilon, 
                        default_alpha=default_alpha, 
                        total_loan=total_loan)
    
def single_asset():
    # default values and init
    ticker = ''
    plot_div = ''
    default_sigma = 0
    default_delta = 10 / 250
    default_epsilon = 0.01
    default_alpha = 0.25
    default_price = 0.00
    sigma, delta, epsilon, alpha, x, adtv, price = default_sigma, default_delta, default_epsilon, default_alpha, 1, 1, default_price
    mu, gamma, lv, liq_lv, loan_amount, error_message = 0, 0, 0, 0, 0, ''

    try:
        # inputs
        ticker = request.form.get('ticker', '').strip().upper()
        x = int(request.form.get('x', 1))
        delta = float(request.form.get('delta', default_delta))
        epsilon = float(request.form.get('epsilon', default_epsilon))
        alpha = float(request.form.get('alpha', default_alpha))
        values_calculated = False

        if ticker:
            values_calculated = True
            swx_list = pd.read_csv(os.path.join('data', 'external', 'swx_symbol_list.csv'))
            tick_list = list(swx_list['Symbol'])

            if ticker not in tick_list:
                price = default_price
                sigma = default_sigma
                adtv = 1
                error_message = f'<b style="color: red;">WARNING:</b> Invalid SWX symbol, using default values instead.'
                ticker = 'INVALID'
            else:
                price = calculate_avg_close_price(ticker)
                sigma = calculate_volatility(ticker)
                adtv = calculate_adtv(ticker)
        
        else:
            # set to default values if no ticker is provided
            price = float(request.form.get('price', default_price))
            sigma = float(request.form.get('sigma', default_sigma))
            adtv = int(request.form.get('ADTV', 1))
        
        # input errors
        if sigma < 0 or sigma > 1 or x < 0 or adtv < 0 or price < 0:
            raise ValueError('Sigma must be between 0 and 1. X, ADTV, and stock price must be non-negative.')
        elif x == 0:
            raise ValueError('Setting transaction size to zero will calculate non-liquidity adjusted LV.')
        elif adtv == 0:
            raise ValueError('ADTV cannot be zero.')
        elif not 0 <= delta <= 1 or not 0 <= epsilon <= 1 or not 0 <= alpha <= 1:
            raise ValueError('Epsilon, Alpha, and Delta must be between 0 and 1.')
    
        # model calc
        mu = (sigma ** 2) / 2
        gamma = 10 ** -1.87096 * adtv ** -0.794554 # lit values
        # gamma = 10 ** -1.6233878972787608 * adtv ** -1.3063903747604177
        lv = lev.calc_lv(mu, sigma, delta, epsilon, alpha, 0, gamma)
        liq_lv = lev.calc_lv(mu, sigma, delta, epsilon, alpha, x, gamma)

        # outputs and formatting
        loan_amount = format(x * liq_lv * price, ',.2f')
        plot_div = single_plot(sigma, delta, epsilon, alpha, x, adtv, price)
        gamma = format(gamma, '.6e')
        lv = format(lv, '.2%')
        liq_lv = format(liq_lv, '.2%')
        sigma = round(sigma, 2)
        adtv = round(adtv)
        price = round(price, 2)

    except ValueError as e:     
        error_message = f'<b style="color: red;">WARNING:</b> {e}'
    
    pass

    return render_template('single_asset.html', ticker=(ticker if values_calculated else 'None'), plot_div=plot_div,
                       error_message=error_message, mu=mu, gamma=gamma, lv=lv, liq_lv=liq_lv, 
                       loan_amount=loan_amount, 
                       sigma=(sigma if values_calculated else request.form.get('sigma', default_sigma)),
                       delta=(delta if values_calculated else request.form.get('delta', default_delta)),
                       epsilon=(epsilon if values_calculated else request.form.get('epsilon', default_epsilon)),
                       alpha=(alpha if values_calculated else request.form.get('alpha', default_alpha)),
                       x=(x if values_calculated else request.form.get('x', 1)),
                       adtv=(adtv if values_calculated else request.form.get('ADTV', 1)),
                       price=(price if values_calculated else request.form.get('price', default_price)),
                       default_sigma=default_sigma, default_delta=default_delta,
                       default_epsilon=default_epsilon, default_alpha=default_alpha, default_price=default_price)